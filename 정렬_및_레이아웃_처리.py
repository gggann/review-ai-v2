"""
정렬_및_레이아웃_처리.py — 검수 완료 후 엑셀 파일 정렬 및 레이아웃 처리
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
재검수 및 회원분류 처리 파일의 서식을 참고하여 작성되었습니다.

[처리 내용]
  1. 랭킹 순으로 데이터 정렬
  2. 특정 열의 너비 축소 (구매확정일, 주문번호, 상품번호, 리뷰타입, 사진, 텍스트, 
     대/중카테고리, 도움돼요, 메인탭등록여부)
  3. 오늘 날짜 폴더에 결과 파일 저장

[출력 폴더 구조]
  output/
  └── YYYYMMDD/
      └── 리뷰동영상검수_YYYYMMDD.xlsx

실행: python 정렬_및_레이아웃_처리.py
또는 main.py에서 import하여 사용
"""

import os
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════
# 설정
# ═══════════════════════════════════════════════════════════════════════════

def get_config():
    """설정값 반환"""
    today_str = datetime.now().strftime('%Y%m%d')
    
    return {
        'BASE_DIR': os.path.dirname(os.path.abspath(__file__)),
        'today_str': today_str,
        'OUTPUT_DIR': os.path.join(os.path.dirname(os.path.abspath(__file__)), "output", today_str),
    }


# ═══════════════════════════════════════════════════════════════════════════
# 열 레이아웃 설정
# ═══════════════════════════════════════════════════════════════════════════

def get_column_width_config():
    """
    열 너비 설정 반환 (사용자 지정 열 번호 기준)
    
    Returns:
        dict: {열인덱스: 너비} 형식의 딕셔너리
              1 = 숨김 처리
    """
    return {
        # 사용자 지정 열 번호 - 너비 1 (숨김 처리)
        2: 1,    # B열: 구매확정일
        3: 1,    # C열: 주문번호
        6: 1,    # F열: 상품번호
        7: 1,    # G열: 리뷰타입
        9: 1,    # I열: 사진
        11: 1,   # K열: 텍스트
        12: 1,   # L열: 대카테고리
        13: 1,   # M열: 중카테고리
        15: 1,   # O열: 도움돼요
        18: 1,   # R열: 메인탭등록여부
    }


# ═══════════════════════════════════════════════════════════════════════════
# 열 인덱스 찾기
# ═══════════════════════════════════════════════════════════════════════════

def find_column_index(ws, column_names):
    """
    워크시트에서 지정된 이름의 열 인덱스를 찾음
    
    Args:
        ws: 워크시트 객체
        column_names: 찾을 열 이름 목록 (리스트)
    
    Returns:
        int: 열 인덱스 (찾지 못하면 None)
    """
    for row_idx in [1, 2]:  # 1행과 2행에서 검색
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).strip()
                for name in column_names:
                    if name in cell_str:
                        return col_idx
    return None


# ═══════════════════════════════════════════════════════════════════════════
# 랭킹 열 찾기
# ═══════════════════════════════════════════════════════════════════════════

def find_ranking_column(ws):
    """
    워크시트에서 랭킹 열 인덱스를 찾음
    
    Returns:
        int: 랭킹 열 인덱스 (찾지 못하면 None)
    """
    ranking_names = ['랭킹', '순위', '-ranking', 'rank']
    return find_column_index(ws, ranking_names)


# ═══════════════════════════════════════════════════════════════════════════
# 열 너비 설정 적용
# ═══════════════════════════════════════════════════════════════════════════

def apply_column_widths(ws, width_config):
    """
    엑셀 워크시트에 열 너비 설정 적용
    
    Args:
        ws: 워크시트 객체
        width_config: {열인덱스: 너비} 딕셔너리
    """
    for col_idx, width in width_config.items():
        if col_idx <= ws.max_column:
            col_letter = get_column_letter(col_idx)
            if width == 0:
                # 숨김 처리
                ws.column_dimensions[col_letter].hidden = True
            else:
                ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════════════════════
# 랭킹 순 정렬
# ═══════════════════════════════════════════════════════════════════════════

def sort_by_ranking(ws, ranking_col_idx, header_offset=2):
    """
    랭킹 열을 기준으로 오름차순 정렬
    
    Args:
        ws: 워크시트 객체
        ranking_col_idx: 랭킹 열 인덱스
        header_offset: 헤더 행 수 (기본값 2)
    """
    if ranking_col_idx is None:
        print("[경고] 랭킹 열을 찾지 못하여 정렬 건너뜀")
        return
    
    # 마지막 행 찾기
    last_row = ws.max_row
    last_col = ws.max_column
    
    if last_row <= header_offset:
        print("[정보] 데이터 행이 없어 정렬 건너뜀")
        return
    
    # 데이터 범위 설정 (헤더 제외)
    # pandas DataFrame으로 정렬
    try:
        df = pd.DataFrame(ws.values)
        
        # 헤더 행 처리
        if header_offset == 2:
            headers = df.iloc[0:2].apply(lambda x: x.astype(str).replace('None', '')).values.tolist()
            data_df = df.iloc[2:].copy()
        else:
            headers = [df.iloc[0].values.tolist()]
            data_df = df.iloc[1:].copy()
        
        if data_df.empty:
            print("[정보] 정렬할 데이터가 없습니다.")
            return
        
        # 랭킹 열 인덱스 (0-based)
        ranking_col_0based = ranking_col_idx - 1
        
        if ranking_col_0based < data_df.shape[1]:
            # 숫자 변환 후 정렬
            data_df[ranking_col_0based] = pd.to_numeric(data_df[ranking_col_0based], errors='coerce')
            data_df = data_df.sort_values(by=ranking_col_0based, ascending=True, na_position='last')
            
            # 정렬된 데이터 다시 쓰기
            for i, row in enumerate(data_df.values):
                for j, value in enumerate(row):
                    ws.cell(row=header_offset + 1 + i, column=j + 1).value = value
            
            print(f"[정보] 랭킹 순 정렬 완료: {len(data_df)}행")
        else:
            print(f"[경고] 랭킹 열 인덱스({ranking_col_idx})가 데이터 범위를 벗어남")
            
    except Exception as e:
        print(f"[오류] 정렬 중 오류 발생: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# J열 텍스트 2 강조 표시
# ═══════════════════════════════════════════════════════════════════════════

from openpyxl.styles import PatternFill, Font, Color

def highlight_text_2_cells(ws, text_col_idx=10, header_offset=2):
    """
    J열(텍스트)에서 값이 2인 셀만 강조 색상으로 표시
    
    Args:
        ws: 워크시트 객체
        text_col_idx: 텍스트 열 인덱스 (기본값 10 = J열)
        header_offset: 헤더 행 수 (기본값 2)
    
    Returns:
        int: 강조된 셀 수
    """
    # 강조 색상 (노란색 배경) - ARGB 형식
    highlight_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    
    highlighted_count = 0
    
    for row_idx in range(header_offset + 1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=text_col_idx).value
        try:
            # 값이 2인지 확인 (숫자 또는 문자열)
            if cell_value is not None:
                cell_str = str(cell_value).strip()
                if cell_str == "2":
                    # J열 셀 자체에만 강조 색상 적용
                    ws.cell(row=row_idx, column=text_col_idx).fill = highlight_fill
                    highlighted_count += 1
        except Exception as e:
            pass
    
    if highlighted_count > 0:
        print(f"[정보] 텍스트 2 강조: {highlighted_count}개 셀 처리 완료")
    
    return highlighted_count


# ═══════════════════════════════════════════════════════════════════════════
# 텍스트 줄 바꿈 설정
# ═══════════════════════════════════════════════════════════════════════════

def apply_text_wrapping(ws, columns=None):
    """
    지정된 열에 텍스트 줄 바꿈 적용
    
    Args:
        ws: 워크시트 객체
        columns: 줄 바꿈을 적용할 열 인덱스 목록 (None이면 모든 열)
    """
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    if columns is None:
        columns = range(1, ws.max_column + 1)
    
    for col_idx in columns:
        if col_idx <= ws.max_column:
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = wrap_alignment


# ═══════════════════════════════════════════════════════════════════════════
# 메인 처리 함수
# ═══════════════════════════════════════════════════════════════════════════

def process_file(input_path, output_path=None, sort_enabled=True, layout_enabled=True):
    """
    엑셀 파일 처리 메인 함수
    
    Args:
        input_path: 입력 파일 경로
        output_path: 출력 파일 경로 (None이면 input_path에 덮어쓰기)
        sort_enabled: 랭킹 정렬 여부 (기본값 True)
        layout_enabled: 레이아웃 적용 여부 (기본값 True)
    
    Returns:
        bool: 처리 성공 여부
    """
    if not os.path.exists(input_path):
        print(f"[오류] 입력 파일 없음: {input_path}")
        return False
    
    print(f"[시작] 파일 처리: {input_path}")
    start_time = time.time()
    
    try:
        # 엑셀 파일 읽기
        wb = load_workbook(input_path)
        ws = wb.active
        
        print(f"[정보] 워크시트: {ws.title}, 행: {ws.max_row}, 열: {ws.max_column}")
        
        # 1. 랭킹 열 찾기
        ranking_col_idx = find_ranking_column(ws)
        if ranking_col_idx:
            print(f"[정보] 랭킹 열 발견: {get_column_letter(ranking_col_idx)}열 ({ranking_col_idx})")
        else:
            print("[경고] 랭킹 열을 찾지 못했습니다. 정렬 건너뜀.")
        
        # 2. 랭킹 순 정렬
        if sort_enabled and ranking_col_idx:
            sort_by_ranking(ws, ranking_col_idx)
        
        # 3. 열 너비 설정 적용
        if layout_enabled:
            width_config = get_column_width_config()
            apply_column_widths(ws, width_config)
            print(f"[정보] 열 너비 설정 적용 완료")
        
        # 4. J열 텍스트 2 강조 표시
        highlight_text_2_cells(ws, text_col_idx=10, header_offset=2)
        
        # 5. 출력 파일 저장
        if output_path is None:
            output_path = input_path
        
        # 출력 디렉토리 생성
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        
        wb.save(output_path)
        wb.close()
        
        elapsed = time.time() - start_time
        print(f"[완료] 파일 저장: {output_path}")
        print(f"[완료] 처리 시간: {elapsed:.2f}초")
        
        return True
        
    except PermissionError:
        print(f"[오류] 파일이 열려있습니다. 파일을 닫고 다시 시도하세요.")
        return False
    except Exception as e:
        print(f"[오류] 처리 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False


def process_directory(input_dir, output_dir=None, sort_enabled=True, layout_enabled=True):
    """
    디렉토리 내 모든 엑셀 파일 처리
    
    Args:
        input_dir: 입력 디렉토리 경로
        output_dir: 출력 디렉토리 경로 (None이면 input_dir과 동일)
        sort_enabled: 랭킹 정렬 여부
        layout_enabled: 레이아웃 적용 여부
    
    Returns:
        dict: 처리 결과 {"success": [], "failed": []}
    """
    config = get_config()
    
    if output_dir is None:
        output_dir = config['OUTPUT_DIR']
    
    results = {"success": [], "failed": []}
    
    # 입력 디렉토리 내 엑셀 파일 검색
    excel_extensions = ['.xlsx', '.xls', '.xlsm']
    excel_files = []
    
    for ext in excel_extensions:
        excel_files.extend(Path(input_dir).glob(f"*{ext}"))
        excel_files.extend(Path(input_dir).glob(f"*{ext.upper()}"))
    
    excel_files = list(set(excel_files))  # 중복 제거
    
    if not excel_files:
        print(f"[경고] 입력 디렉토리에 엑셀 파일이 없습니다: {input_dir}")
        return results
    
    print(f"[정보] 발견된 파일: {len(excel_files)}개")
    
    for file_path in excel_files:
        # 출력 경로 설정
        relative_path = file_path.relative_to(input_dir) if input_dir != file_path.parent else file_path.name
        output_path = Path(output_dir) / relative_path
        
        success = process_file(
            str(file_path), 
            str(output_path),
            sort_enabled=sort_enabled,
            layout_enabled=layout_enabled
        )
        
        if success:
            results["success"].append(str(file_path))
        else:
            results["failed"].append(str(file_path))
    
    # 결과 요약
    print(f"\n{'='*60}")
    print(f"처리 결과 요약")
    print(f"{'='*60}")
    print(f"  성공: {len(results['success'])}개")
    print(f"  실패: {len(results['failed'])}개")
    print(f"  출력 폴더: {output_dir}")
    print(f"{'='*60}")
    
    return results


# ═══════════════════════════════════════════════════════════════════════════
# 테스트용 샘플 데이터 생성
# ═══════════════════════════════════════════════════════════════════════════

def create_sample_excel(output_path):
    """
    테스트용 샘플 엑셀 파일 생성
    
    Args:
        output_path: 출력 파일 경로
    """
    config = get_config()
    
    # 샘플 데이터
    data = {
        '구매확정일': ['2026-04-13', '2026-04-13', '2026-04-13', '2026-04-13', '2026-04-13'],
        '주문번호': ['ORD001', 'ORD002', 'ORD003', 'ORD004', 'ORD005'],
        '상품번호': ['P1001', 'P1002', 'P1003', 'P1004', 'P1005'],
        '리뷰타입': ['동영상', '동영상', '텍스트', '동영상', '텍스트'],
        '사진': ['5', '3', '0', '7', '2'],
        '텍스트': ['좋은 제품입니다', '빠른 배송', '가격 대비 좋은', '추천합니다', '만족'],
        '대카테고리': ['전자기기', '의류', '식품', '전자기기', '식품'],
        '중카테고리': ['스마트폰', '여성의류', '건강식품', '태블릿', '과일'],
        '도움돼요': ['10', '5', '2', '15', '3'],
        '랭킹': [3, 1, 5, 2, 4],
        '기타열1': ['데이터1', '데이터2', '데이터3', '데이터4', '데이터5'],
        '기타열2': ['값1', '값2', '값3', '값4', '값5'],
        '메인탭등록여부': ['검수자A', '검수자B', '검수자A', '검수자C', '검수자B'],
    }
    
    df = pd.DataFrame(data)
    
    # 출력 디렉토리 생성
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # 엑셀로 저장
    df.to_excel(output_path, index=False, sheet_name='검수결과')
    
    print(f"[샘플] 테스트 파일 생성 완료: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# 진입점
# ═══════════════════════════════════════════════════════════════════════════

def main():
    """메인 진입점"""
    config = get_config()
    
    print(f"{'='*60}")
    print(f"정렬 및 레이아웃 처리 도구")
    print(f"{'='*60}")
    print(f"실행 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"오늘 날짜: {config['today_str']}")
    print(f"기본 출력: {config['OUTPUT_DIR']}")
    print(f"{'='*60}")
    
    # 테스트용 샘플 파일 생성 및 처리
    sample_input = os.path.join(config['BASE_DIR'], "sample_input.xlsx")
    sample_output = os.path.join(config['OUTPUT_DIR'], "sample_output.xlsx")
    
    # 샘플 파일 생성
    create_sample_excel(sample_input)
    
    # 샘플 파일 처리
    print()
    success = process_file(
        sample_input,
        sample_output,
        sort_enabled=True,
        layout_enabled=True
    )
    
    if success:
        print(f"\n[SUCCESS] 테스트 완료!")
        print(f"출력 파일: {sample_output}")
    else:
        print(f"\n[FAILED] 테스트 실패!")


if __name__ == "__main__":
    main()
