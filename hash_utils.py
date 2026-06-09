"""
hash_utils.py — 하이브리드 중복 영상 감지 (바이너리 해시 + 프레임 해시)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
3단계 하이브리드 전략: 속도 + 정확도 동시 확보

[처리 흐름]
  ├─ 1단계: 파일 크기 그룹핑 (즉시)
  ├─ 2단계: Quick MD5 (파일 앞 1MB, OpenCV 없음)
  └─ 3단계: 프레임 해시 (2단계 통과 + 동일 회원만)

[중간 저장]
  · _hash_cache.json: 해시 결과 캐시 (재실행 시 스킵)
"""

import os
import json
import hashlib
import asyncio
import aiohttp
import time
import concurrent.futures
from collections import defaultdict
from datetime import datetime

import cv2
import imagehash
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from config import (
    log,
    get_video_path,
    DUPLICATE_HASH_THRESHOLD,
    DUPLICATE_DURATION_MARGIN,
    TEMP_VIDEO_DIR,
    HASH_CACHE_PATH,
    QUICK_HASH_BYTES,
    FRAME_SAMPLE_POSITIONS,
    HASH_FRAME_SIZE,
    CPU_WORKERS,
    HASH_REPORT_DIR,
    MAX_CONCURRENT_TASKS,
)
from video_utils import download_video_direct, extract_cont_no


# ═══════════════════════════════════════════════════════════════════════════
# 1단계 + 2단계: 파일 크기 + Quick MD5 (OpenCV 없음, 초고속)
# ═══════════════════════════════════════════════════════════════════════════

def compute_quick_fingerprint(video_path: str) -> dict | None:
    """
    바이너리 지문 계산 (OpenCV 미사용, 극히 빠름).
    · file_size: 전체 파일 크기
    · quick_md5: 앞 1MB의 MD5 해시
    """
    try:
        if not os.path.exists(video_path):
            return None
        file_size = os.path.getsize(video_path)
        if file_size < 1000:
            return None
        with open(video_path, "rb") as f:
            data = f.read(QUICK_HASH_BYTES)
            quick_md5 = hashlib.md5(data).hexdigest()
        return {
            "file_size": file_size,
            "quick_md5": quick_md5,
        }
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════
# 3단계: 프레임 해시 (OpenCV — ProcessPoolExecutor에서 실행)
# ═══════════════════════════════════════════════════════════════════════════

def compute_frame_hash(video_path: str) -> dict | None:
    """
    CPU bound 프레임 해시 계산.
    고정 위치 3프레임(10%/50%/90%) + 128×128 리사이즈.
    """
    try:
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            return None

        fc = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        duration = (fc / fps) if fps > 0 else 0.0

        if fc <= 0 or fps <= 0 or duration < 0.5:
            cap.release()
            return None

        avg_hashes = []
        p_hashes = []

        for ratio in FRAME_SAMPLE_POSITIONS:
            target = max(0, min(int(fc * ratio), fc - 1))
            cap.set(cv2.CAP_PROP_POS_FRAMES, target)
            ret, frame = cap.read()
            if not ret:
                continue
            small = cv2.resize(frame, HASH_FRAME_SIZE)
            pil_img = Image.fromarray(cv2.cvtColor(small, cv2.COLOR_BGR2RGB))
            avg_hashes.append(str(imagehash.average_hash(pil_img)))
            p_hashes.append(str(imagehash.phash(pil_img)))

        cap.release()

        if not avg_hashes:
            return None

        return {
            "avg_hashes": avg_hashes,
            "p_hashes": p_hashes,
            "duration_sec": round(duration, 2),
        }
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════
# 해시 유사도 비교
# ═══════════════════════════════════════════════════════════════════════════

def _hamming(h1_str: str, h2_str: str) -> int:
    return imagehash.hex_to_hash(h1_str) - imagehash.hex_to_hash(h2_str)


def are_frames_similar(rec1: dict, rec2: dict) -> bool:
    """프레임 해시 기반 시각적 유사도 판정"""
    if abs(rec1["duration_sec"] - rec2["duration_sec"]) > DUPLICATE_DURATION_MARGIN:
        return False
    avg1, avg2 = rec1["avg_hashes"], rec2["avg_hashes"]
    ph1, ph2 = rec1["p_hashes"], rec2["p_hashes"]
    pairs = min(len(avg1), len(avg2))
    if pairs == 0:
        return False
    avg_dist = sum(_hamming(avg1[k], avg2[k]) for k in range(pairs)) / pairs
    if avg_dist > DUPLICATE_HASH_THRESHOLD:
        return False
    ph_dist = sum(_hamming(ph1[k], ph2[k]) for k in range(pairs)) / pairs
    return ph_dist <= DUPLICATE_HASH_THRESHOLD * 2


# ═══════════════════════════════════════════════════════════════════════════
# Union-Find
# ═══════════════════════════════════════════════════════════════════════════

class UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))

    def find(self, x):
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, x, y):
        self.parent[self.find(x)] = self.find(y)

    def groups(self):
        g = defaultdict(list)
        for i in range(len(self.parent)):
            g[self.find(i)].append(i)
        return list(g.values())


# ═══════════════════════════════════════════════════════════════════════════
# 캐시
# ═══════════════════════════════════════════════════════════════════════════

def load_hash_cache() -> dict:
    if os.path.exists(HASH_CACHE_PATH):
        try:
            with open(HASH_CACHE_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def save_hash_cache(cache: dict):
    try:
        os.makedirs(os.path.dirname(HASH_CACHE_PATH), exist_ok=True)
        with open(HASH_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f"[경고] 해시 캐시 저장 실패: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# 영상 경로 확정
# ═══════════════════════════════════════════════════════════════════════════

async def _resolve_video_path(session, row_data, sem):
    """contNo 기반 LAN URL 다운로드"""
    i, rev_id, h_url, category_info, text_len, member_id = row_data

    h_url_str = str(h_url) if h_url and str(h_url) != "nan" else ""
    cont_no = extract_cont_no(h_url_str) if h_url_str.startswith("http") else None

    # contNo 기반 LAN URL만 사용
    async with sem:
        urls_to_try = []
        if cont_no:
            urls_to_try.append(get_video_path(cont_no))

        if not urls_to_try:
            return {
                "index": i, "review_id": str(rev_id), "member_id": str(member_id),
                "video_path": "", "is_temp": False, "success": False,
                "category_info": category_info,
            }

        tmp = os.path.join(TEMP_VIDEO_DIR, f"_hash_{i}_{rev_id}.mp4")
        for url in urls_to_try:
            if await download_video_direct(session, url, tmp):
                return {
                    "index": i, "review_id": str(rev_id), "member_id": str(member_id),
                    "video_path": tmp, "is_temp": True, "success": True,
                    "category_info": category_info,
                }

        return {
            "index": i, "review_id": str(rev_id), "member_id": str(member_id),
            "video_path": "", "is_temp": False, "success": False,
            "category_info": category_info,
        }


# ═══════════════════════════════════════════════════════════════════════════
# 메인: 하이브리드 중복 감지
# ═══════════════════════════════════════════════════════════════════════════

async def precompute_video_groups(
    session,
    tasks_data: list,
    temp_dir: str,
    max_concurrent: int = 20,
    representative_strategy: str = "first",
) -> tuple[dict, list]:
    """
    3단계 하이브리드 중복 감지.

    Returns:
        (video_groups, all_records)
        · video_groups: {review_id: {is_representative, original_review_id, ...}}
        · all_records:  [{review_id, member_id, file_size, quick_md5, avg_hashes, ...}]
    """
    t_start = time.time()
    stage_times = {}

    # ── Step 0: 영상 경로 확정 ──
    t0 = time.time()
    log(f"\n[Phase1] 영상 경로 확정 — {len(tasks_data)}건")
    download_sem = asyncio.Semaphore(max_concurrent)
    path_tasks = [_resolve_video_path(session, td, download_sem) for td in tasks_data]

    try:
        from tqdm.asyncio import tqdm_asyncio
        path_results = await tqdm_asyncio.gather(*path_tasks, desc="[경로 확정]")
    except Exception:
        path_results = await asyncio.gather(*path_tasks)

    local_count = sum(1 for r in path_results if r["success"] and not r["is_temp"])
    dl_count = sum(1 for r in path_results if r["success"] and r["is_temp"])
    fail_count = sum(1 for r in path_results if not r["success"])
    log(f"[Phase1] 경로 확정 완료 — 로컬={local_count} / 다운로드={dl_count} / 실패={fail_count}")
    stage_times["resolve"] = time.time() - t0

    pr_map = {pr["review_id"]: pr for pr in path_results}

    # ── Step 1+2: 바이너리 지문 (파일 크기 + Quick MD5) ──
    t1 = time.time()
    log(f"\n[Phase1] 바이너리 지문 계산 (파일크기 + Quick MD5)")

    all_records = []
    cache = load_hash_cache()

    from tqdm import tqdm
    for pr in tqdm(path_results, desc="[바이너리 해시]"):
        rev_id = pr["review_id"]
        rec = {
            "review_id": rev_id,
            "member_id": pr["member_id"],
            "index": pr["index"],
            "category_info": pr.get("category_info", ""),
            "video_path": pr.get("video_path", ""),
            "is_temp": pr.get("is_temp", False),
            "success": False,
            "file_size": 0,
            "quick_md5": "",
            "avg_hashes": [],
            "p_hashes": [],
            "duration_sec": 0.0,
        }

        if not pr["success"] or not pr["video_path"]:
            all_records.append(rec)
            continue

        # 캐시 확인 (바이너리 + 프레임 해시 모두)
        if rev_id in cache:
            cached = cache[rev_id]
            rec["file_size"] = cached.get("file_size", 0)
            rec["quick_md5"] = cached.get("quick_md5", "")
            rec["avg_hashes"] = cached.get("avg_hashes", [])
            rec["p_hashes"] = cached.get("p_hashes", [])
            rec["duration_sec"] = cached.get("duration_sec", 0.0)
            rec["success"] = True
            all_records.append(rec)
            continue

        # Quick Fingerprint 계산
        fp = compute_quick_fingerprint(pr["video_path"])
        if fp:
            rec["file_size"] = fp["file_size"]
            rec["quick_md5"] = fp["quick_md5"]
            rec["success"] = True
        all_records.append(rec)

    binary_ok = sum(1 for r in all_records if r["success"])
    log(f"[Phase1] 바이너리 지문 완료 — 성공={binary_ok} / 실패={len(all_records) - binary_ok}")

    # ── 바이너리 중복 찾기 ──
    member_groups = defaultdict(list)
    for rec in all_records:
        if rec["success"]:
            member_groups[rec["member_id"]].append(rec)

    binary_dup_pairs = set()
    binary_dup_count = 0

    for member_id, recs in member_groups.items():
        if len(recs) < 2:
            continue
        sig_map = defaultdict(list)
        for rec in recs:
            sig = f"{rec['file_size']}_{rec['quick_md5']}"
            sig_map[sig].append(rec)

        for sig, group in sig_map.items():
            if len(group) > 1:
                for a in range(len(group)):
                    for b in range(a + 1, len(group)):
                        pair = tuple(sorted([group[a]["review_id"], group[b]["review_id"]]))
                        binary_dup_pairs.add(pair)
                binary_dup_count += len(group) - 1

    log(f"[Phase1] 바이너리 중복: {binary_dup_count}건 (OpenCV 스킵)")
    stage_times["binary"] = time.time() - t1

    # ── Step 3: 프레임 해시 (같은 회원, 바이너리로 못 잡은 것만) ──
    t2 = time.time()

    need_frame_hash = set()
    for member_id, recs in member_groups.items():
        if len(recs) < 2:
            continue
        for rec in recs:
            # 이미 프레임 해시가 캐시에 있으면 스킵
            if rec.get("avg_hashes"):
                continue
            # 같은 회원 내 바이너리로 매칭 안 된 쌍이 있으면 프레임 해시 필요
            for other in recs:
                if other["review_id"] == rec["review_id"]:
                    continue
                pair = tuple(sorted([rec["review_id"], other["review_id"]]))
                if pair not in binary_dup_pairs:
                    need_frame_hash.add(rec["review_id"])
                    break

    need_opencv = []
    for rev_id in need_frame_hash:
        rec = next((r for r in all_records if r["review_id"] == rev_id), None)
        if rec and rec["success"] and rec["video_path"]:
            need_opencv.append((rev_id, rec["video_path"]))

    log(f"\n[Phase1] 프레임 해시 (OpenCV) — {len(need_opencv)}건 "
        f"({len(need_opencv)/max(len(tasks_data),1)*100:.1f}%만 정밀 검사)")

    if need_opencv:
        opencv_results = {}
        with concurrent.futures.ProcessPoolExecutor(max_workers=CPU_WORKERS) as executor:
            futures_map = {}
            for rev_id, video_path in need_opencv:
                future = executor.submit(compute_frame_hash, video_path)
                futures_map[future] = rev_id

            pbar = tqdm(total=len(futures_map), desc="[프레임 해시(CPU)]")
            for future in concurrent.futures.as_completed(futures_map):
                rev_id = futures_map[future]
                try:
                    result = future.result()
                except Exception:
                    result = None
                opencv_results[rev_id] = result
                pbar.update(1)
            pbar.close()

        for rev_id, result in opencv_results.items():
            rec = next((r for r in all_records if r["review_id"] == rev_id), None)
            if rec and result:
                rec["avg_hashes"] = result["avg_hashes"]
                rec["p_hashes"] = result["p_hashes"]
                rec["duration_sec"] = result["duration_sec"]
                # 캐시 업데이트
                if rev_id not in cache:
                    cache[rev_id] = {}
                cache[rev_id]["avg_hashes"] = result["avg_hashes"]
                cache[rev_id]["p_hashes"] = result["p_hashes"]
                cache[rev_id]["duration_sec"] = result["duration_sec"]

    stage_times["visual"] = time.time() - t2

    # ── 캐시에 바이너리 정보 저장 ──
    for rec in all_records:
        if rec["success"] and rec["review_id"] not in cache:
            cache[rec["review_id"]] = {
                "file_size": rec["file_size"],
                "quick_md5": rec["quick_md5"],
                "success": True,
            }
        elif rec["success"] and rec["review_id"] in cache:
            cache[rec["review_id"]].setdefault("file_size", rec["file_size"])
            cache[rec["review_id"]].setdefault("quick_md5", rec["quick_md5"])

    save_hash_cache(cache)

    # ── Step 4: 통합 클러스터링 (바이너리 + 시각적) ──
    t3 = time.time()
    log(f"\n[Phase1] 통합 클러스터링")

    video_groups = {}
    group_counter = 0

    for member_id, recs in member_groups.items():
        if len(recs) == 1:
            video_groups[recs[0]["review_id"]] = {
                "is_representative": True,
                "original_review_id": None,
                "member_id": member_id,
                "group_id": group_counter,
                "group_review_ids": [recs[0]["review_id"]],
                "match_type": "—",
            }
            group_counter += 1
            continue

        n = len(recs)
        uf = UnionFind(n)
        match_types = {}

        for a in range(n):
            for b in range(a + 1, n):
                pair = tuple(sorted([recs[a]["review_id"], recs[b]["review_id"]]))
                if pair in binary_dup_pairs:
                    uf.union(a, b)
                    match_types[(a, b)] = "binary"
                    continue
                if recs[a].get("avg_hashes") and recs[b].get("avg_hashes"):
                    if are_frames_similar(recs[a], recs[b]):
                        uf.union(a, b)
                        match_types[(a, b)] = "visual"

        clusters = uf.groups()
        for cluster_indices in clusters:
            cluster_recs = [recs[k] for k in cluster_indices]

            # 대표 선정
            if representative_strategy == "latest":
                best = max(cluster_recs, key=lambda r: r["review_id"])
            elif representative_strategy == "longest_video":
                best = max(cluster_recs, key=lambda r: r.get("duration_sec", 0))
            else:
                best = min(cluster_recs, key=lambda r: r["index"])
            rep_id = best["review_id"]
            group_review_ids = [r["review_id"] for r in cluster_recs]

            # 그룹 매치 유형
            group_match = "—"
            if len(cluster_recs) > 1:
                has_binary = any(
                    match_types.get((min(a, b), max(a, b))) == "binary"
                    for a in cluster_indices for b in cluster_indices if a < b
                )
                has_visual = any(
                    match_types.get((min(a, b), max(a, b))) == "visual"
                    for a in cluster_indices for b in cluster_indices if a < b
                )
                if has_binary and has_visual:
                    group_match = "mixed"
                elif has_binary:
                    group_match = "binary"
                elif has_visual:
                    group_match = "visual"

            for rec in cluster_recs:
                is_rep = rec["review_id"] == rep_id
                video_groups[rec["review_id"]] = {
                    "is_representative": is_rep,
                    "original_review_id": None if is_rep else rep_id,
                    "member_id": member_id,
                    "group_id": group_counter,
                    "group_review_ids": group_review_ids,
                    "match_type": group_match if is_rep else (
                        "binary" if tuple(sorted([rec["review_id"], rep_id])) in binary_dup_pairs
                        else "visual"
                    ),
                }

            if len(cluster_recs) > 1:
                dup_ids = [r["review_id"] for r in cluster_recs if r["review_id"] != rep_id]
                log(f"[Phase1] 클러스터 — member={member_id} group={group_counter} "
                    f"유형={group_match} 대표={rep_id} 중복={len(dup_ids)}건")

            group_counter += 1

    stage_times["cluster"] = time.time() - t3

    total_rep = sum(1 for v in video_groups.values() if v["is_representative"])
    total_dup = sum(1 for v in video_groups.values() if not v["is_representative"])
    total_time = time.time() - t_start
    log(f"[Phase1] 완료 — 대표={total_rep} 중복={total_dup} "
        f"({total_time:.1f}초: 경로={stage_times.get('resolve',0):.1f}s "
        f"바이너리={stage_times.get('binary',0):.1f}s "
        f"프레임={stage_times.get('visual',0):.1f}s "
        f"클러스터={stage_times.get('cluster',0):.1f}s)")

    # ── 임시 파일 정리 ──
    for rec in all_records:
        if rec.get("is_temp") and rec.get("video_path") and os.path.exists(rec["video_path"]):
            try:
                os.remove(rec["video_path"])
            except Exception:
                pass

    return video_groups, all_records


# ═══════════════════════════════════════════════════════════════════════════
# Excel 보고서 (과도기용)
# ═══════════════════════════════════════════════════════════════════════════

FILL_DUP_BIN = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
FILL_DUP_VIS = PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid")
FILL_REP = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
FILL_SINGLE = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
FILL_FAIL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
FILL_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
FONT_BOLD = Font(bold=True)
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")


def save_hash_report(all_records: list, video_groups: dict, tasks_data: list):
    """
    하이브리드 중복 감지 결과를 엑셀로 저장 (hash_report/날짜_중복체크결과.xlsx).
    """
    os.makedirs(HASH_REPORT_DIR, exist_ok=True)
    today_str = datetime.now().strftime('%Y%m%d')
    output_path = os.path.join(HASH_REPORT_DIR, f"{today_str}_중복체크결과.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "하이브리드_중복체크"

    headers = [
        "No", "리뷰번호", "회원번호", "카테고리",
        "파일크기(KB)", "Quick_MD5", "영상길이(초)",
        "중복유형", "중복여부", "대표리뷰", "그룹ID",
        "그룹내리뷰목록", "avg_hash", "phash"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rec_map = {r["review_id"]: r for r in all_records}

    row_num = 2
    stats = {"binary_dup": 0, "visual_dup": 0, "fail": 0, "single": 0, "rep_group": 0}

    for td in tasks_data:
        i, rev_id, h_url, category_info, text_len, member_id = td
        rec = rec_map.get(str(rev_id), {})
        grp = video_groups.get(str(rev_id), {})

        file_kb = round(rec.get("file_size", 0) / 1024, 1) if rec.get("file_size") else 0
        quick_md5 = rec.get("quick_md5", "")
        duration = rec.get("duration_sec", 0.0)

        if not rec.get("success", False):
            dup_type, dup_label = "—", "해시실패"
            stats["fail"] += 1
        elif not grp:
            dup_type, dup_label = "—", "해시실패"
            stats["fail"] += 1
        elif grp.get("is_representative", True):
            group_ids = grp.get("group_review_ids", [])
            if len(group_ids) > 1:
                dup_type = grp.get("match_type", "—")
                dup_label = "대표 (중복그룹)"
                stats["rep_group"] += 1
            else:
                dup_type, dup_label = "—", "단독"
                stats["single"] += 1
        else:
            match_type = grp.get("match_type", "visual")
            dup_type = "바이너리(100%)" if match_type == "binary" else "시각적(프레임)"
            dup_label = "중복"
            stats["binary_dup" if match_type == "binary" else "visual_dup"] += 1

        original_rep = grp.get("original_review_id", "") or ""
        group_id = grp.get("group_id", "") if grp else ""
        group_list = ", ".join(grp.get("group_review_ids", [])) if grp else ""
        avg_hash_str = " | ".join(rec.get("avg_hashes", []))
        phash_str = " | ".join(rec.get("p_hashes", []))

        row_values = [
            i + 1, str(rev_id), str(member_id), category_info,
            file_kb, quick_md5[:12] + "..." if len(quick_md5) > 12 else quick_md5,
            round(duration, 1),
            dup_type, dup_label, original_rep, group_id,
            group_list, avg_hash_str, phash_str
        ]
        for col, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = ALIGN_WRAP

        # 색상
        if dup_label == "중복" and dup_type.startswith("바이너리"):
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = FILL_DUP_BIN
                if c in (8, 9):
                    ws.cell(row=row_num, column=c).font = FONT_WHITE_BOLD
        elif dup_label == "중복":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = FILL_DUP_VIS
        elif dup_label == "대표 (중복그룹)":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = FILL_REP
        elif dup_label == "해시실패":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = FILL_FAIL

        row_num += 1

    # 통계
    row_num += 1
    total = len(tasks_data)
    summary = [
        ("═══ 하이브리드 통계 ═══", ""),
        ("전체 대상", f"{total}건"),
        ("바이너리 중복 (100% 동일)", f"{stats['binary_dup']}건"),
        ("시각적 중복 (프레임 유사)", f"{stats['visual_dup']}건"),
        ("중복 합계", f"{stats['binary_dup'] + stats['visual_dup']}건"),
        ("해시 실패", f"{stats['fail']}건"),
    ]
    for label, val in summary:
        ws.cell(row=row_num, column=1, value=label).font = FONT_BOLD
        c = ws.cell(row=row_num, column=2, value=val)
        if "바이너리" in label:
            c.fill = FILL_DUP_BIN
            c.font = FONT_WHITE_BOLD
        elif "시각적" in label:
            c.fill = FILL_DUP_VIS
        row_num += 1

    # 열 너비
    widths = {1: 6, 2: 14, 3: 14, 4: 25, 5: 12, 6: 15, 7: 12,
              8: 18, 9: 16, 10: 14, 11: 8, 12: 40, 13: 30, 14: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    log(f"\n[Phase1] 해시 보고서 저장: {output_path}")
    return stats
