"""
test_region_mse.py — 정지영상 판별 (3프레임 중심부 MSE)
엑셀 출력: 정지영상_영역별MSE분석.xlsx
"""
import os, sys, io
import numpy as np
import cv2
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SRC_DIR = r"E:\nginx-1.20.2\dist\리뷰 AI 정식버전2\정지영상모음"
OUT_XLSX = os.path.join(os.path.dirname(SRC_DIR), "정지영상_영역별MSE분석.xlsx")
THRESHOLD = 30


def _mse(a, b):
    a = a.flatten().astype("float")
    b = b.flatten().astype("float")
    return np.mean((a - b) ** 2)


def analyze(vp: str) -> dict | None:
    cap = cv2.VideoCapture(vp)
    if not cap.isOpened():
        return None

    fc = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    fps = cap.get(cv2.CAP_PROP_FPS)
    dur = (fc / fps) if fps > 0 else 0.0

    if fc <= 0 or fps <= 0:
        cap.release()
        return None

    # 3프레임: 처음, 중간, 끝
    mid = fc // 2
    cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
    ret1, f1 = cap.read()
    cap.set(cv2.CAP_PROP_POS_FRAMES, mid)
    ret_m, fm = cap.read()
    cap.set(cv2.CAP_PROP_POS_FRAMES, fc - 1)
    ret2, f2 = cap.read()
    cap.release()

    if not (ret1 and ret2):
        return None

    g1 = cv2.cvtColor(f1, cv2.COLOR_BGR2GRAY)
    g2 = cv2.cvtColor(f2, cv2.COLOR_BGR2GRAY)
    g_m = cv2.cvtColor(fm, cv2.COLOR_BGR2GRAY) if ret_m else None
    if g1.shape != g2.shape:
        h, w = g1.shape
        g2 = cv2.resize(g2, (w, h))
    if g_m is not None and g1.shape != g_m.shape:
        h, w = g1.shape
        g_m = cv2.resize(g_m, (w, h))

    # 중심부(안쪽 40%)만 비교
    h, w = g1.shape
    mh, mw = int(h * 0.30), int(w * 0.30)

    def _cmse(a, b):
        return _mse(a[mh:h - mh, mw:w - mw], b[mh:h - mh, mw:w - mw])

    first_last = _cmse(g1, g2)
    center_mse = first_last
    if g_m is not None:
        center_mse = max(first_last, _cmse(g1, g_m), _cmse(g_m, g2))

    return {
        "file": os.path.basename(vp),
        "duration": round(dur, 2),
        "frames": fc,
        "center_mse": round(center_mse, 1),
    }


def main():
    files = sorted(f for f in os.listdir(SRC_DIR) if f.endswith(".mp4"))
    print(f"대상: {len(files)}개")

    results = []
    for f in files:
        r = analyze(os.path.join(SRC_DIR, f))
        if r:
            results.append(r)

    results.sort(key=lambda x: x["center_mse"])

    wb = Workbook()
    ws = wb.active
    ws.title = "3프레임중심부MSE"

    headers = ["No", "파일명(링크)", "길이(초)", "총프레임", "중심부MSE", "판정"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center")

    fill_static = PatternFill("solid", fgColor="90EE90")
    fill_pass = PatternFill("solid", fgColor="FFCC66")

    for i, r in enumerate(results):
        row = i + 2
        judge = "Static" if r["center_mse"] < THRESHOLD else "Pass"
        cont_no = r["file"].replace(".mp4", "")
        link = f"http://10.240.129.13/review/{cont_no}/{cont_no}.mp4"

        vals = [i + 1, link, r["duration"], r["frames"], r["center_mse"], judge]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)

        for c in (5, 6):
            ws.cell(row=row, column=c).fill = fill_static if judge == "Static" else fill_pass

    # 통계
    row = len(results) + 3
    ws.cell(row=row, column=1, value="Static 판정 건수").font = Font(bold=True)
    cnt = sum(1 for r in results if r["center_mse"] < THRESHOLD)
    ws.cell(row=row + 1, column=1, value=f"MSE < {THRESHOLD}")
    ws.cell(row=row + 1, column=2, value=f"{cnt}개")

    widths = [5, 50, 10, 10, 12, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    wb.save(OUT_XLSX)
    print(f"\n저장: {OUT_XLSX}")
    print("=" * 60)
    for r in results:
        j = "S" if r["center_mse"] < 100 else "P"
        print(f"  {r['file']:20s}  중심부={r['center_mse']:>7.1f}({j})")


if __name__ == "__main__":
    main()
