"""
main.py - 리뷰 영상 AI 검수 시스템
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
엑셀 파일의 리뷰 데이터를 읽어 AI 기반으로 채점합니다.

[처리 흐름]
  1. 엑셀 → 리뷰 데이터 읽기
  2. 영상 다운로드 + 프레임 추출
  3. 1단계: 부적합 영상 필터링
  4. 2단계: 10점제 평가 (품질/가시성/행동성)
  5. 포인트 지급 점수 산출
  6. 엑셀 → 결과 저장

[출력 컬럼 순서]
  포인트 지급 점수 → 10점_총점 → 품질 → 가시성 → 행동성
  → 판정단계 → 판정사유 → 상세사유
  → 부적합_판정 → 부적합_사유 → 부적합_감지항목
  → 추출_옵션명 → 추출_리뷰본문

실행: python main.py
"""

import os
import sys
import json
import time
import asyncio
import socket
import traceback
import concurrent.futures
from datetime import datetime, timedelta
from glob import glob

import pandas as pd
import aiohttp
from tqdm.asyncio import tqdm_asyncio
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from config import (
    INPUT_EXCEL_PATH, OUTPUT_FILE, OUTPUT_DIR, TEMP_VIDEO_DIR, BASE_DIR,
    GPT_MODEL, MAX_CONCURRENT_TASKS, TEST_MODE, TEST_LIMIT,
    _CONFIG_INFO, get_video_path,
    log, load_zero_point_rules, openai_client,
)
from 정렬_및_레이아웃_처리 import (
    process_file, find_ranking_column, sort_by_ranking,
    apply_column_widths, get_column_width_config, highlight_text_2_cells
)
from video_utils import (
    get_image_b64,
    download_video_direct, extract_frames,
    load_local_metadata, extract_cont_no,
)
from hash_utils import (
    precompute_video_groups,
    save_hash_report,
)
from zero_point import run_zero_point_phase
from five_point import run_5point_phase


# ═══════════════════════════════════════════════════════════════════════════
# 설정 & 상수
# ═══════════════════════════════════════════════════════════════════════════

# 평가 등급 순위 (높은 값 = 좋은 결과)
GRADE_RANK = {"10": 10, "9": 9, "8": 8, "7": 7, "6": 6,
              "5": 5, "4": 4, "3": 3, "2": 2, "1": 1, "0": 0,
              "Check": -1, "Error": -2}

# ─── 엑셀 색상 팔레트 ───
class ExcelStyles:
    """엑셀 셀 스타일 모음"""
    
    # 점수별 배경색
    FILL_SCORE_HIGH = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 8-10점: 녹색
    FILL_SCORE_MID  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 5-7점: 노랑
    FILL_SCORE_LOW  = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 1-4점: 금색
    FILL_SCORE_ZERO = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 0점: 빨강
    
    # 포인트 지급 배경색
    FILL_PT_5  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 5점: 녹색
    FILL_PT_2  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 2점: 노랑
    FILL_PT_0  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 0점: 연한빨강
    
    # 부적합 관련
    FILL_ZERO  = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")  # 부적합(0점)
    FILL_DUP   = PatternFill(start_color="E8A0BF", end_color="E8A0BF", fill_type="solid")  # 중복리뷰(연한자주색)
    FILL_PASS  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 통과
    
    # 폰트
    FONT_BOLD  = Font(bold=True)
    FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
    
    # 정렬
    WRAP_TOP = Alignment(wrap_text=True, vertical="top")


# ─── 영상 상태 메시지 ───
VIDEO_STATUS_MESSAGES = {
    "ZeroSecond":  lambda d: f"비정상 영상 (길이 {d:.1f}초 / 1초 미만)",
    "StaticImage": lambda d: f"비정상 영상 (정지 이미지 / {d:.1f}초)",
    "BlackScreen": "비정상 영상 (검은 화면)",
    "Error":       "비정상 영상 (프레임 추출 실패)"
}


# ═══════════════════════════════════════════════════════════════════════════
# 점수 계산
# ═══════════════════════════════════════════════════════════════════════════

def calc_point_score(total_score: int, action_score: int, zero_score: str) -> int:
    """
    포인트 지급 점수 계산 (0 / 2 / 5점 체계)
    
    규칙:
      • 1단계 부적합(zero_score == "0") → 0점 (포인트 미지급)
      • 2단계 총점 7~10점              → 5점
      • 2단계 총점 5~6점 + 행동성 ≥ 2  → 5점
      • 그 외                           → 2점
    """
    if str(zero_score) == "0":
        return 0

    # Check/Error 등 비정상 점수 → 포인트 미지급
    if str(total_score) in ("Check", "Error", ""):
        return 0

    try:
        ts = int(total_score)
        ac = int(action_score)
    except (TypeError, ValueError):
        return 2

    if ts >= 7:
        return 5
    if ts >= 5 and ac >= 2:
        return 5
    return 2


def get_score_fill(total_score: int) -> PatternFill:
    """총점에 따른 배경색 반환"""
    if total_score >= 8:
        return ExcelStyles.FILL_SCORE_HIGH
    if 5 <= total_score <= 7:
        return ExcelStyles.FILL_SCORE_MID
    if 1 <= total_score <= 4:
        return ExcelStyles.FILL_SCORE_LOW
    return ExcelStyles.FILL_SCORE_ZERO


def get_point_fill(pt_score: int) -> tuple[PatternFill, Font]:
    """포인트 점수에 따른 배경색+폰트 반환"""
    if pt_score == 5:
        return ExcelStyles.FILL_PT_5, ExcelStyles.FONT_BOLD
    if pt_score == 2:
        return ExcelStyles.FILL_PT_2, ExcelStyles.FONT_BOLD
    return ExcelStyles.FILL_PT_0, ExcelStyles.FONT_WHITE_BOLD


# ═══════════════════════════════════════════════════════════════════════════
# 중복 그룹 후처리
# ═══════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════
# 리뷰 처리
# ═══════════════════════════════════════════════════════════════════════════

async def process_one_review(sem: asyncio.Semaphore, session: aiohttp.ClientSession,
                              row_data: tuple, zero_rules: dict,
                              video_groups: dict = None) -> dict:
    """
    리뷰 1건 통합 처리 파이프라인

    단계별 처리:
      [0] Phase 1 중복 영상 → 즉시 스킵 (API 호출 없음)
      [1] 메타데이터 수집 (로컬)
      [2] 영상 다운로드 (로컬/LAN 경로)
      [3] 프레임 추출
      [4] 부적합 탐지 (1단계 필터)
      [5] 10점제 평가 (2단계, 부적합 pass 건만)
      
    변경사항: 중복 영상은 Phase 1에서 즉시 스킵 (API 호출 없이 중복리뷰 마킹).
    """
    async with sem:
        i, rev_id, h_url, category_info, text_len, member_id = row_data
        t0 = time.time()

        # ── 기본값 초기화 ──
        default = {
            "index": i, "review_id": rev_id,
            "mp4_url": "", "opt_name": "", "review_text": "",
            "zero_score": "", "zero_reason": "", "zero_features": [],
            "quality": 0, "visibility": 0, "action": 0,
            "total_score": 0, "quality_reason": "",
            "visibility_reason": "", "action_reason": "",
            "overall_reason": "", "applied_category_rules": [],
            "reason": "유효한 링크 없음", "eval_stage": ""
        }

        # ═══════════════════════════════════════════
        # [0] Phase 1 중복 영상 → 즉시 스킵
        # ═══════════════════════════════════════════
        is_rep_of_dup_group = False
        if video_groups:
            group_info = video_groups.get(str(rev_id))
            if group_info:
                if not group_info.get("is_representative", True):
                    rep_id = group_info.get("original_review_id", "")
                    log(f"[리뷰 {rev_id}] [0]중복 영상 스킵 (대표: {rep_id})")
                    return {
                        **default,
                        "zero_score": "0",
                        "zero_reason": f"중복 영상 감지 (대표 리뷰: {rep_id})",
                        "zero_features": ["duplicate_video"],
                        "total_score": 0,
                        "reason": f"중복 영상 감지 (대표 리뷰: {rep_id})",
                        "eval_stage": "중복스킵",
                        "duplicate_flag": "중복리뷰",
                    }
                elif len(group_info.get("group_review_ids", [])) > 1:
                    is_rep_of_dup_group = True

        # ═══════════════════════════════════════════
        # [1] 메타데이터 수집
        # ═══════════════════════════════════════════
        mp4_links, p_img_b64, p_name, p_opt, p_text = [], "", "", "", ""

        if pd.notna(h_url) and str(h_url).startswith("http"):
            t1 = time.time()
            cont_no = extract_cont_no(str(h_url))
            local_meta = await load_local_metadata(session, cont_no) if cont_no else None

            if local_meta:
                # 로컬 메타데이터에서 정보 추출
                p_name = local_meta.get("product", {}).get("name", "")
                p_opt  = local_meta.get("product", {}).get("option", "")
                p_text = local_meta.get("review",  {}).get("text", "")
                meta_status = local_meta.get("status", {})

                if meta_status.get("image_ok"):
                    image_url = local_meta.get("product", {}).get("image_local_url", "")
                else:
                    image_url = local_meta.get("product", {}).get("image_source_url", "")
                if image_url:
                    p_img_b64 = await get_image_b64(session, image_url)

                # 영상: 로컬 경로 우선
                if cont_no:
                    mp4_links.append(get_video_path(cont_no))

                log(f"[리뷰 {rev_id}] [1]로컬메타 {time.time()-t1:.1f}s | mp4={len(mp4_links)}개")

            default["opt_name"]    = p_opt
            default["review_text"] = p_text

        if not mp4_links:
            log(f"[리뷰 {rev_id}] mp4 링크 없음 → 스킵 ({time.time()-t0:.1f}s)")
            return default

        # ═══════════════════════════════════════════
        # [2-5] 영상 처리 (LAN URL 1건)
        # ═══════════════════════════════════════════
        url = mp4_links[0]
        url_tag = "로컬" if os.path.isabs(url) else "LAN"
        tmp = os.path.join(TEMP_VIDEO_DIR, f"integrated_{i}_0.mp4")

        # ── [2] 다운로드 ──
        t2 = time.time()
        dl_ok = await download_video_direct(session, url, tmp)
        log(f"[리뷰 {rev_id}] [2]다운로드({url_tag}) {time.time()-t2:.1f}s | "
            f"{'성공' if dl_ok else '실패'}")

        if not dl_ok:
            if os.path.exists(tmp):
                os.remove(tmp)
            result = default
        else:
            # ── [3] 프레임 추출 ──
            t3 = time.time()
            status, frames_bgr, duration_sec, avg_motion = extract_frames(tmp)
            log(f"[리뷰 {rev_id}] [3]프레임추출 {time.time()-t3:.1f}s | "
                f"status={status} frames={len(frames_bgr) if frames_bgr else 0}")

            if os.path.exists(tmp):
                os.remove(tmp)

            # 프레임 추출 실패 시 → 영상 오류 처리
            if status != "Pass":
                reason = VIDEO_STATUS_MESSAGES.get(status, f"비정상 ({status})")
                if callable(reason):
                    reason = reason(duration_sec)
                result = {
                    **default,
                    "mp4_url": url,
                    "zero_score": "0",
                    "zero_reason": reason,
                    "total_score": 0,
                    "duration_sec": duration_sec,
                    "reason": reason,
                    "eval_stage": "영상(오류)",
                }
            else:
                # ── [4] 부적합 탐지 (1단계) ──
                t4 = time.time()
                zero_res = await run_zero_point_phase(
                    frames_bgr, duration_sec, p_img_b64, p_name, p_opt,
                    category_info, str(rev_id), str(member_id), zero_rules
                )
                log(f"[리뷰 {rev_id}] [4]부적합탐지 {time.time()-t4:.1f}s | zero={zero_res['zero_score']}")

                if zero_res["zero_score"] == "0":
                    result = {
                        **default,
                        "mp4_url": url,
                        "opt_name": p_opt,
                        "review_text": p_text,
                        "zero_score": zero_res["zero_score"],
                        "zero_reason": zero_res["zero_reason"],
                        "zero_features": zero_res["zero_features"],
                        "total_score": 0,
                        "duration_sec": duration_sec,
                        "reason": zero_res["zero_reason"],
                        "eval_stage": "부적합",
                    }
                else:
                    # ── [5] 10점제 평가 (2단계) ──
                    t5 = time.time()
                    quality_res = await run_5point_phase(frames_bgr, category_info)
                    score = quality_res.get("total_score", 0)
                    stage = quality_res.get("eval_stage", "?")
                    log(f"[리뷰 {rev_id}] [5]10점제평가 {time.time()-t5:.1f}s | total={score} stage={stage}")

                    reason_dict = quality_res.get("reason", {})
                    result = {
                        **default,
                        "mp4_url": url,
                        "opt_name": p_opt,
                        "review_text": p_text,
                        "zero_score": zero_res["zero_score"],
                        "zero_reason": zero_res["zero_reason"],
                        "zero_features": zero_res["zero_features"],
                        "quality": quality_res.get("quality", 0),
                        "visibility": quality_res.get("visibility", 0),
                        "action": quality_res.get("action", 0),
                        "total_score": score,
                        "quality_reason": reason_dict.get("quality_reason", ""),
                        "visibility_reason": reason_dict.get("visibility_reason", ""),
                        "action_reason": reason_dict.get("action_reason", ""),
                        "overall_reason": reason_dict.get("overall_reason", ""),
                        "applied_category_rules": quality_res.get("applied_category_rules", []),
                        "reason": reason_dict.get("overall_reason", ""),
                        "eval_stage": stage,
                    }

        result["index"] = i
        result["review_id"] = rev_id
        if is_rep_of_dup_group:
            result["duplicate_flag"] = "중복리뷰"
        log(f"[리뷰 {rev_id}] 완료 총 {time.time()-t0:.1f}s | "
            f"최종={result.get('total_score')}점 zero={result.get('zero_score')}")
        return result


async def safe_process(sem: asyncio.Semaphore, session: aiohttp.ClientSession,
                       row_data: tuple, zero_rules: dict,
                       video_groups: dict = None) -> dict:
    """리뷰 처리 중 예외 발생 시 안전하게 처리"""
    try:
        return await process_one_review(sem, session, row_data, zero_rules, video_groups)
    except Exception as e:
        error_detail = traceback.format_exc()
        return {
            "index": row_data[0], "review_id": row_data[1],
            "mp4_url": "", "opt_name": "", "review_text": "",
            "zero_score": "Error", "zero_reason": f"시스템 에러: {str(e)}",
            "zero_features": [],
            "quality": 0, "visibility": 0, "action": 0,
            "total_score": 0, "applied_category_rules": [],
            "reason": f"[에러] {type(e).__name__}: {str(e)}",
            "error_detail": error_detail, "eval_stage": "에러"
        }


# ═══════════════════════════════════════════════════════════════════════════
# 엑셀 처리
# ═══════════════════════════════════════════════════════════════════════════

def prepare_excel_output(ws, results: list, header_offset: int) -> tuple[dict, dict, list]:
    """
    엑셀에 결과写入 및 통계 산출
    
    Returns:
        stats: 10점제 점수 분포
        zero_stats: 부적합 탐지 결과 분포
        scores_list: 총점 리스트 (평균 계산용)
    """
    # ─── 컬럼 인덱스 설정 (U열=21부터 시작) ───
    START_COL = 21
    new_col_counter = [0]

    def get_col(name: str) -> int:
        """기존 헤더에서 찾고, 없으면 START_COL부터 신규 추가"""
        headers = [str(ws.cell(row=header_offset, column=c).value or "").strip()
                   for c in range(1, ws.max_column + 1)]
        if name in headers:
            return headers.index(name) + 1
        new_col_idx = START_COL + new_col_counter[0]
        new_col_counter[0] += 1
        ws.cell(row=header_offset, column=new_col_idx).value = name
        return new_col_idx

    # 컬럼 정의
    C_POINT_SCORE = get_col("포인트 지급 점수")
    C_TOTAL_5PT   = get_col("10점_총점")
    C_QUALITY      = get_col("품질(0-2)")
    C_VISIBIL      = get_col("가시성(0-4)")
    C_ACTION       = get_col("행동성(0-4)")
    C_STAGE        = get_col("판정단계")
    C_REASON       = get_col("판정사유")
    C_DETAIL       = get_col("상세사유")
    C_CAT_RULES    = get_col("적용_카테고리규칙")
    C_ZERO_SCORE   = get_col("부적합_판정")
    C_ZERO_REASON  = get_col("부적합_사유")
    C_ZERO_FEAT    = get_col("부적합_감지항목")
    C_OPT          = get_col("추출_옵션명")
    C_TEXT_EXT     = get_col("추출_리뷰본문")
    C_DUP_FLAG     = get_col("중복여부")

    # ─── 행 크기 조정 (열 너비는 main.py에서 별도 처리) ───
    for row_idx in range(1, ws.max_row + len(results) + 5):
        ws.row_dimensions[row_idx].height = 14.75

    # ─── 통계 수집 ───
    stats = {"10": 0, "9": 0, "8": 0, "7": 0, "6": 0,
             "5": 0, "4": 0, "3": 0, "2": 0, "1": 0, "0": 0,
             "Check": 0, "Error": 0}
    zero_stats = {"0": 0, "pass": 0, "Check": 0, "Error": 0}
    scores_list = []

    # ─── 결과 작성 ───
    for res in results:
        row = res["index"] + header_offset + 1
        total_score = res.get("total_score", 0)
        z_score = res.get("zero_score", "")
        action_sc = res.get("action", 0)
        stage = res.get("eval_stage", "")

        # 통계 업데이트
        if total_score is not None:
            stats[str(total_score)] = stats.get(str(total_score), 0) + 1
            if total_score > 0:
                scores_list.append(total_score)
        else:
            stats[res.get("eval_stage", "Error")] = stats.get(res.get("eval_stage", "Error"), 0) + 1
        zero_stats[z_score] = zero_stats.get(z_score, 0) + 1

        # ── 포인트 지급 점수 ──
        pt_score = calc_point_score(total_score, action_sc, z_score)
        pt_cell = ws.cell(row=row, column=C_POINT_SCORE)
        pt_cell.value = pt_score
        pt_cell.fill, pt_cell.font = get_point_fill(pt_score)

        # ── 10점제 총점 ──
        score_cell = ws.cell(row=row, column=C_TOTAL_5PT)
        score_cell.value = total_score
        score_cell.font = ExcelStyles.FONT_BOLD
        score_cell.fill = get_score_fill(total_score)

        # ── 세부 점수 ──
        ws.cell(row=row, column=C_QUALITY).value = res.get("quality", 0)
        ws.cell(row=row, column=C_VISIBIL).value = res.get("visibility", 0)
        ws.cell(row=row, column=C_ACTION).value = res.get("action", 0)

        # ── 판정 정보 ──
        ws.cell(row=row, column=C_STAGE).value = stage
        ws.cell(row=row, column=C_REASON).value = res.get("reason", "")

        # ── 상세 사유 ──
        detail_parts = []
        if res.get("quality_reason"):
            detail_parts.append(f"[품질] {res['quality_reason']}")
        if res.get("visibility_reason"):
            detail_parts.append(f"[가시성] {res['visibility_reason']}")
        if res.get("action_reason"):
            detail_parts.append(f"[행동성] {res['action_reason']}")
        if res.get("overall_reason"):
            detail_parts.append(f"[종합] {res['overall_reason']}")
        if res.get("error_detail"):
            detail_parts.append(f"【에러】\n{res['error_detail']}")
        d_cell = ws.cell(row=row, column=C_DETAIL)
        d_cell.value = "\n".join(detail_parts)
        d_cell.alignment = ExcelStyles.WRAP_TOP

        # ── 적용 카테고리 규칙 ──
        cat_rules = res.get("applied_category_rules", [])
        cat_cell = ws.cell(row=row, column=C_CAT_RULES)
        cat_cell.value = "\n".join(cat_rules) if cat_rules else ""
        cat_cell.alignment = ExcelStyles.WRAP_TOP

        # ── 부적합 정보 ──
        z_cell = ws.cell(row=row, column=C_ZERO_SCORE)
        z_cell.value = z_score
        z_cell.font = ExcelStyles.FONT_BOLD
        if z_score == "0":
            z_cell.fill = ExcelStyles.FILL_ZERO
            z_cell.font = ExcelStyles.FONT_WHITE_BOLD
        elif z_score == "pass":
            z_cell.fill = ExcelStyles.FILL_PASS

        ws.cell(row=row, column=C_ZERO_REASON).value = res.get("zero_reason", "")
        ws.cell(row=row, column=C_ZERO_FEAT).value = ", ".join(res.get("zero_features", []))

        # ── 추출 텍스트 ──
        if res.get("opt_name"):
            ws.cell(row=row, column=C_OPT).value = res["opt_name"]
        if res.get("review_text"):
            ws.cell(row=row, column=C_TEXT_EXT).value = res["review_text"]

        # ── 중복여부 ──
        dup_flag = res.get("duplicate_flag", "")
        dup_cell = ws.cell(row=row, column=C_DUP_FLAG)
        dup_cell.value = dup_flag
        if dup_flag == "중복리뷰":
            dup_cell.fill = ExcelStyles.FILL_DUP
            dup_cell.font = ExcelStyles.FONT_BOLD

    return stats, zero_stats, scores_list


# ═══════════════════════════════════════════════════════════════════════════
# 결과 출력 포맷
# ═══════════════════════════════════════════════════════════════════════════

def print_summary(total: int, stats: dict, zero_stats: dict, scores_list: list):
    """결과 요약 출력"""
    avg_score = sum(scores_list) / len(scores_list) if scores_list else 0
    
    sep = "=" * 60
    
    print(f"\n{sep}")
    print("📊 10점제 통합 채점 결과")
    print(f"{sep}")
    print(f"   총 검수: {total}건")
    
    print(f"\n  ┌─ 1단계: 부적합 탐지")
    print(f"  │   ├─ 0점 (부적합):    {zero_stats.get('0', 0):>3}건")
    print(f"  │   ├─ pass (정상):     {zero_stats.get('pass', 0):>3}건")
    print(f"  │   └─ 확인 필요:       {zero_stats.get('Check', 0):>3}건")
    
    print(f"\n  ┌─ 2단계: 10점제 평가 (부적합 pass 건)")
    print(f"  │   ├─ 10점 (우수):     {stats.get('10', 0):>3}건  ({stats.get('10', 0)/total*100:>5.1f}%)")
    print(f"  │   ├─ 8-9점 (좋음):   {stats.get('9', 0) + stats.get('8', 0):>3}건  "
          f"({(stats.get('9', 0) + stats.get('8', 0))/total*100:>5.1f}%)")
    print(f"  │   ├─ 6-7점 (보통):   {stats.get('7', 0) + stats.get('6', 0):>3}건  "
          f"({(stats.get('7', 0) + stats.get('6', 0))/total*100:>5.1f}%)")
    print(f"  │   ├─ 4-5점 (미흡):   {stats.get('5', 0) + stats.get('4', 0):>3}건  "
          f"({(stats.get('5', 0) + stats.get('4', 0))/total*100:>5.1f}%)")
    print(f"  │   ├─ 2-3점 (낮음):   {stats.get('3', 0) + stats.get('2', 0):>3}건  "
          f"({(stats.get('3', 0) + stats.get('2', 0))/total*100:>5.1f}%)")
    print(f"  │   ├─ 0-1점 (미달):   {stats.get('1', 0) + stats.get('0', 0):>3}건  "
          f"({(stats.get('1', 0) + stats.get('0', 0))/total*100:>5.1f}%)")
    print(f"  │   ├─ 확인 필요:       {stats.get('Check', 0):>3}건")
    print(f"  │   └─ 에러:           {stats.get('Error', 0):>3}건")
    
    print(f"\n  📈 평균 총점: {avg_score:.1f} / 10점")
    print(f"\n{sep}")
    print(f"💾 결과 저장: {OUTPUT_FILE}")
    print(f"📁 출력 폴더: {OUTPUT_DIR}")
    print(sep)


# ═══════════════════════════════════════════════════════════════════════════
# Preflight API 연결 테스트
# ═══════════════════════════════════════════════════════════════════════════

def preflight_api_check() -> bool:
    """
    중계 서버(API) 연결 사전 테스트.
    실제 업무와 동일한 조건(JSON 응답 + 모델)으로 검증합니다.
    실패 시 False 반환 → 본 처리를 시작하지 않습니다.
    """
    log("[Preflight] 중계 서버 연결 테스트 중...")
    log(f"[Preflight]    서버: {openai_client.base_url} | 모델: {GPT_MODEL}")
    try:
        response = openai_client.chat.completions.create(
            model=GPT_MODEL,
            messages=[
                {"role": "system", "content": "JSON으로만 응답하세요."},
                {"role": "user", "content": '1+1은? {"answer": 숫자} 형식으로 답하세요.'}
            ],
            response_format={"type": "json_object"},
            max_completion_tokens=20,
            temperature=0.0,
        )
        if response and response.choices:
            content = response.choices[0].message.content.strip()
            log(f"[Preflight] ✅ 연결 성공 - 모델 응답 확인: {content}")
            return True
        else:
            log("[Preflight] ❌ 중계 서버 응답이 비정상입니다 (choices 없음).")
            return False
    except Exception as e:
        log(f"[Preflight] ❌ 중계 서버 연결 실패: {type(e).__name__}: {e}")
        log(f"[Preflight]    서버 URL: {openai_client.base_url}")
        log(f"[Preflight]    서버가 닫혀있거나 네트워크 문제를 확인하세요.")
        return False


# ═══════════════════════════════════════════════════════════════════════════
# 메인 처리
# ═══════════════════════════════════════════════════════════════════════════

async def main_process():
    """메인 처리: 엑셀 읽기 → 비동기 처리 → 엑셀 저장"""
    
    # ─── 입력 파일 확인 ───
    if not os.path.exists(INPUT_EXCEL_PATH):
        log(f"[ERROR] 입력 파일 없음: {INPUT_EXCEL_PATH}")
        return

    # ─── Preflight: 중계 서버 연결 테스트 ───
    if not preflight_api_check():
        log("[중단] 중계 서버에 연결할 수 없어 작업을 시작하지 않습니다.")
        return

    # ─── 규칙 로드 ───
    zero_rules = load_zero_point_rules()
    log(f"[INFO] zero_point_rules 버전: {zero_rules.get('_version', '?')}")

    # ─── .xls → .xlsx 변환 (가장 먼저 수행) ───
    input_path = INPUT_EXCEL_PATH
    temp_xlsx_path = None
    
    if INPUT_EXCEL_PATH.lower().endswith(".xls"):
        # 이전 변환 파일 정리
        for old_xlsx in glob(os.path.join(BASE_DIR, "리뷰동영상*.xlsx")):
            try:
                os.remove(old_xlsx)
                log(f"[INFO] 이전 변환 파일 삭제: {os.path.basename(old_xlsx)}")
            except Exception:
                pass
        
        xls_basename = os.path.splitext(os.path.basename(INPUT_EXCEL_PATH))[0]
        temp_xlsx_path = os.path.join(BASE_DIR, xls_basename + ".xlsx")
        try:
            temp_wb_df = pd.read_excel(INPUT_EXCEL_PATH, header=None)
            temp_wb_df.to_excel(temp_xlsx_path, index=False, header=False)
            log(f"[INFO] .xls → .xlsx 변환 완료")
            input_path = temp_xlsx_path
        except Exception as conv_err:
            log(f"[오류] .xls → .xlsx 변환 실패: {conv_err}")
            return

    # ─── 엑셀 파일 열기 및 레이아웃 처리 (AI 검수 전) ───
    try:
        wb = load_workbook(input_path)
        ws = wb.active
    except PermissionError:
        log("[오류] 입력 파일이 열려 있습니다.")
        return

    # ─── A열 날짜 읽어서 출력 경로 동적 생성 ───
    try:
        from config import AI_OUTPUT_DIR
        from datetime import timedelta
        from collections import Counter
        
        # A열에서 모든 날짜 수집 (헤더 제외)
        date_values = []
        date_patterns = ['%Y/%m/%d', '%Y-%m-%d', '%Y.%m.%d', '%Y%m%d']
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            
            if cell_value is None:
                continue
            
            # datetime 객체인 경우
            if isinstance(cell_value, datetime):
                date_values.append(cell_value.strftime('%Y%m%d'))
                continue
            
            # 문자열 파싱
            cell_str = str(cell_value).strip()
            for fmt in date_patterns:
                try:
                    parsed_date = datetime.strptime(cell_str, fmt)
                    date_values.append(parsed_date.strftime('%Y%m%d'))
                    break
                except ValueError:
                    continue
        
        # 가장 빈번한 날짜(최빈값) 선택
        file_date = None
        if date_values:
            date_counter = Counter(date_values)
            file_date = date_counter.most_common(1)[0][0]
            log(f"[INFO] [날짜] A열 날짜 분포: {dict(date_counter)}")
            log(f"[INFO] [날짜] 최빈 날짜({file_date}) 선택")
        else:
            # 날짜를 찾지 못하면 어제 날짜 사용
            yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y%m%d')
            file_date = yesterday
            log(f"[경고] A열에서 날짜를 찾지 못함, 어제 날짜({yesterday})를 기본값으로 사용")
        
        # 동적으로 OUTPUT_DIR과 OUTPUT_FILE 재설정
        # 파일명 형식: {날짜}_리뷰동영상검수.xlsx
        import config as _cfg
        _cfg.OUTPUT_DIR = os.path.join(AI_OUTPUT_DIR, file_date)
        _cfg.OUTPUT_FILE = os.path.join(_cfg.OUTPUT_DIR, f"{file_date}_리뷰동영상검수.xlsx")
        
        # main.py에서 참조하기 위해 로컬 변수도 갱신
        global OUTPUT_DIR, OUTPUT_FILE
        OUTPUT_DIR = _cfg.OUTPUT_DIR
        OUTPUT_FILE = _cfg.OUTPUT_FILE
        
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        log(f"[INFO] [날짜] 출력 날짜({file_date}) 기준으로 출력 경로 설정")
        log(f"[INFO] [날짜] 출력 폴더: {OUTPUT_DIR}")
        log(f"[INFO] [날짜] 출력 파일: {OUTPUT_FILE}")
    except Exception as date_err:
        log(f"[경고] 날짜 읽기 오류: {date_err}")

    # ─── 헤더 형식 확인 (데이터 읽기 전에 수행) ───
    temp_df = pd.read_excel(input_path, nrows=2, header=None)
    if "동영상 리뷰 목록" in str(temp_df.iloc[0, 0]):
        df = pd.read_excel(input_path, header=1)
        header_offset = 2
    else:
        df = pd.read_excel(input_path, header=0)
        header_offset = 1
    df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
    
    log(f"[DEBUG] 컬럼: {df.columns.tolist()}")
    log(f"[DEBUG] 총 데이터: {len(df)}행")

    # ─── AI 검수 전: 랭킹 순 정렬 ───
    ranking_col_idx = find_ranking_column(ws)
    if ranking_col_idx:
        log(f"[INFO] [사전] 랭킹 열 발견: 열 {ranking_col_idx}")
        sort_by_ranking(ws, ranking_col_idx, header_offset)
        # 정렬 후 ws를 임시 저장 (pandas가 읽을 수 있도록)
        wb.save(input_path)
        log(f"[INFO] [사전] 랭킹 순 정렬 후 파일 저장 완료")
        # 정렬 후 데이터 다시 읽기
        temp_df = pd.read_excel(input_path, nrows=2, header=None)
        if "동영상 리뷰 목록" in str(temp_df.iloc[0, 0]):
            df = pd.read_excel(input_path, header=1)
            header_offset = 2
        else:
            df = pd.read_excel(input_path, header=0)
            header_offset = 1
        df.columns = df.columns.astype(str).str.strip().str.replace('\n', '')
        log(f"[INFO] [사전] 랭킹 순 정렬 후 데이터 재로드: {len(df)}행")
    else:
        log(f"[INFO] [사전] 랭킹 열을 찾지 못하여 정렬 건너뜀")

    # ─── AI 검수 전: 열 너비 조절 적용 ───
    try:
        width_config = get_column_width_config()
        apply_column_widths(ws, width_config)
        log(f"[INFO] [사전] 열 너비 설정 적용 완료")
    except Exception as layout_err:
        log(f"[경고] 열 너비 설정 중 오류: {layout_err}")

    # ─── 처리 대상 추출 ───
    def get_idx(names, default):
        for n in names:
            if n in df.columns:
                return df.columns.get_loc(n)
        return default

    tasks_data = []
    for i in range(len(df)):
        vid_flag = df.iloc[i, get_idx(["동영상"], 9)]
        if str(vid_flag).strip() not in ("1", "2"):
            continue
        h_url = df.iloc[i, get_idx(["리뷰URL", "URL"], 7)]
        if pd.isna(h_url) or not str(h_url).startswith("http"):
            continue
        rev_id = df.iloc[i, get_idx(["리뷰번호"], 3)]
        mem_id = df.iloc[i, get_idx(["회원번호", "고객번호"], 4)]

        text_len_raw = df.iloc[i, get_idx(["텍스트"], 10)]
        try:
            text_len = int(text_len_raw) if text_len_raw and str(text_len_raw) != 'nan' else 0
        except:
            text_len = 0

        cat1 = str(df.iloc[i, get_idx(["대카테고리"], 11)]).strip()
        cat2 = str(df.iloc[i, get_idx(["중카테고리"], 12)]).strip()
        cat3 = str(df.iloc[i, get_idx(["소카테고리"], 13)]).strip()
        cat_parts = [c for c in [cat1, cat2, cat3] if c and c.lower() != "nan"]
        category_info = " > ".join(cat_parts)

        tasks_data.append((
            i, str(rev_id), str(h_url),
            category_info, text_len,
            str(mem_id) if pd.notna(mem_id) else ""
        ))

    if not tasks_data:
        log("[알림] 검수할 데이터가 없습니다.")
        return
    if TEST_MODE:
        tasks_data = tasks_data[:TEST_LIMIT]

    # ─── 비동기 처리 실행 ───
    log(f"\n[🚀 통합 채점 START] 동영상 대상 {len(tasks_data)}건")
    log(f"   모델: {GPT_MODEL}  |  동시 처리: {MAX_CONCURRENT_TASKS}건")
    log("─" * 60)

    executor = concurrent.futures.ThreadPoolExecutor(max_workers=MAX_CONCURRENT_TASKS * 2)
    loop = asyncio.get_event_loop()
    loop.set_default_executor(executor)

    connector = aiohttp.TCPConnector(
        limit=50, use_dns_cache=True, ttl_dns_cache=600,
        resolver=aiohttp.ThreadedResolver(), family=socket.AF_INET,
    )
    async with aiohttp.ClientSession(connector=connector) as session:

        # ── Phase 1: 하이브리드 중복 감지 ──
        log(f"\n[Phase 1] 하이브리드 중복 감지 시작...")
        video_groups, hash_records = await precompute_video_groups(
            session=session,
            tasks_data=tasks_data,
            temp_dir=TEMP_VIDEO_DIR,
            max_concurrent=MAX_CONCURRENT_TASKS,
            representative_strategy="first",
        )
        n_rep = sum(1 for v in video_groups.values() if v["is_representative"])
        n_dup = sum(1 for v in video_groups.values() if not v["is_representative"])
        n_unresolved = len(tasks_data) - len(video_groups)
        log(f"[Phase 1] 결과 - 대표 리뷰: {n_rep}건 / 중복 건너뜀: {n_dup}건 / 해시실패: {n_unresolved}건")
        log("─" * 60)

        # ── 해시 보고서 저장 (과도기용) ──
        try:
            save_hash_report(hash_records, video_groups, tasks_data)
        except Exception as e:
            log(f"[경고] 해시 보고서 저장 실패: {e}")

        # ── Phase 2: AI 평가 (체크포인트 지원) ──
        CHECKPOINT_INTERVAL = 100  # 100건마다 중간 저장
        checkpoint_path = os.path.join(OUTPUT_DIR, "_checkpoint.json")

        # 체크포인트 로드: 이미 처리된 결과 복원
        results = []
        completed_ids = set()
        remaining_data = list(tasks_data)  # 아직 처리하지 않은 항목

        if os.path.exists(checkpoint_path):
            try:
                with open(checkpoint_path, "r", encoding="utf-8") as f:
                    ckpt = json.load(f)
                completed_ids = set(ckpt.get("completed_ids", []))
                # 체크포인트에 저장된 결과도 복원
                for r in ckpt.get("results", []):
                    results.append(r)
                # remaining_data에서 완료된 항목 제거
                remaining_data = [d for d in tasks_data if str(d[1]) not in completed_ids]
                log(f"[체크포인트] {len(completed_ids)}건 복원, {len(remaining_data)}건 남음")
            except Exception as e:
                log(f"[경고] 체크포인트 로드 실패, 처음부터 시작: {e}")
                results = []
                completed_ids = set()
                remaining_data = list(tasks_data)

        if remaining_data:
            total_count = len(tasks_data)
            log(f"\n[Phase 2] AI 평가 시작 — {len(remaining_data)}건 (전체 {total_count}건 중 {len(completed_ids)}건 완료)")
            log("─" * 60)

            sem = asyncio.Semaphore(MAX_CONCURRENT_TASKS)
            tasks = [safe_process(sem, session, d, zero_rules, video_groups) for d in remaining_data]

            from tqdm import tqdm
            pbar = tqdm(total=len(remaining_data), desc="[AI 채점]")
            for coro in asyncio.as_completed(tasks):
                result = await coro
                results.append(result)
                completed_ids.add(str(result.get("review_id", "")))
                pbar.update(1)

                # CHECKPOINT_INTERVAL마다 중간 저장
                if len(results) % CHECKPOINT_INTERVAL == 0:
                    # 체크포인트 JSON 저장
                    try:
                        ckpt_data = {
                            "completed_ids": list(completed_ids),
                            "results": results,
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                        with open(checkpoint_path, "w", encoding="utf-8") as f:
                            json.dump(ckpt_data, f, ensure_ascii=False)
                    except Exception as e:
                        log(f"[경고] 체크포인트 저장 실패: {e}")

                    # 중간 xlsx 저장
                    try:
                        prepare_excel_output(ws, list(results), header_offset)
                        wb.save(OUTPUT_FILE)
                        log(f"[체크포인트] {len(results)}건 중간 저장 완료")
                    except Exception as e:
                        log(f"[경고] 중간 xlsx 저장 실패: {e}")

            pbar.close()
        else:
            log(f"\n[Phase 2] 모든 AI 평가가 이미 완료됨 ({len(results)}건)")

    # ─── AI 검수 후: 결과 기록 ───
    # (wb, ws는 이미 앞에서 열려 있음)
    stats, zero_stats, scores_list = prepare_excel_output(ws, results, header_offset)

    # ─── AI 검수 후: 열 너비 재적용 및 텍스트 2 강조 ───
    try:
        width_config = get_column_width_config()
        apply_column_widths(ws, width_config)
        log(f"[INFO] [사후] 열 너비 재적용 완료")
        # J열 텍스트 2 강조
        highlighted = highlight_text_2_cells(ws, text_col_idx=10, header_offset=header_offset)
        log(f"[INFO] [사후] 텍스트 2 강조: {highlighted}개 셀 처리")
    except Exception as layout_err:
        log(f"[경고] 레이아웃 처리 중 오류: {layout_err}")

    try:
        wb.save(OUTPUT_FILE)
    except PermissionError:
        log("[오류] 출력 파일이 열려 있습니다.")
        return

    # ─── 완료 후 체크포인트 삭제 ───
    _ckpt = os.path.join(OUTPUT_DIR, "_checkpoint.json")
    if os.path.exists(_ckpt):
        os.remove(_ckpt)

    # ─── 결과 출력 ───
    print_summary(len(results), stats, zero_stats, scores_list)


# ═══════════════════════════════════════════════════════════════════════════
# 진입점
# ═══════════════════════════════════════════════════════════════════════════

def main():
    # config.py에서 모은 시작 메시지 출력
    for msg in _CONFIG_INFO:
        print(msg)

    print(f"[시작] 10점제 통합 AI 검수 시작")
    print(f"       입력 파일: {INPUT_EXCEL_PATH}")

    if not os.path.exists(INPUT_EXCEL_PATH):
        print(f"\n[오류] 입력 파일을 찾을 수 없습니다!")
        print(f"       경로: {INPUT_EXCEL_PATH}")
        return

    try:
        asyncio.run(main_process())
    except Exception as e:
        print(f"[오류] 처리 중 오류 발생: {e}")

    print("[완료] 모든 작업 완료")


if __name__ == "__main__":
    main()